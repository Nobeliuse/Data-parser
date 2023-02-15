from selenium import webdriver
from bs4 import BeautifulSoup
import time
import copy
from openpyxl import load_workbook, Workbook


class SaverExcel:
	@staticmethod
	def __get_excel_file():
		wb = load_workbook('data.xlsx')
		ws = wb.active
		return wb, ws

	@staticmethod
	def __get_array_names_match(ws):
		excel_data = []
		for cell in ws['A']:
			excel_data.append(str(cell.value))
		return excel_data

	def save_data_to_excel_line(self, data_line):
		try:
			wb, ws = self.__get_excel_file()
			excel_data = self.__get_array_names_match(ws)
		except FileNotFoundError:
			name_columns = [
				"The match date and time",
				"The match's name",
				"The liga's name",
				"w1",
				"Draw",
				"w2",
				"Total",
				'More total',
				'Less total',
				'Total Live',
				'More total live',
				'Less total live',
				'Status'
			] 
			wb = Workbook() 
			ws = wb.active
			i = 0
			ws.append(name_column for name_column in name_columns)
			while i != len(data_line['name']):
				ws.append(data_line[data][i] for data in data_line if data_line[data] != [])
				i += 1
		finally:
			wb.save('data.xlsx')

	def save_data_to_excel_live(self, data_live):
		while True:
			wb, ws = self.__get_excel_file()
			excel_data = self.__get_array_names_match(ws)
			i = 0
			if data_live == None:
				print('No match at the moment')
				return None	
			new_sheet = wb.get_sheet_by_name('Sheet')
			while i != len(data_live['name']):
				for cell in ws['A']:
					if cell.value == data_live['name'][i]:
						new_sheet.cell(row=cell.row, column=4).value = data_live['w1'][i]
						new_sheet.cell(row=cell.row, column=5).value = data_live['draw'][i]
						new_sheet.cell(row=cell.row, column=6).value = data_live['w2'][i]
						new_sheet.cell(row=cell.row, column=10).value = data_live['total_live'][i]
						new_sheet.cell(row=cell.row, column=11).value = data_live['mtotal_live'][i]
						new_sheet.cell(row=cell.row, column=12).value = data_live['ltotal_live'][i]
				i += 1
			wb.save('data.xlsx')


class ParserData(SaverExcel):
	data_parser = {
		'name': [],
		'time_data': [],
		'liga_name': [],
		'w1': [],
		'draw': [],
		'w2': [],
		'total': [],
		'ltotal': [],
		'mtotal': [],
		'status': [],
		'total_live': [],
		'mtotal_live':[],
		'ltotal_live':[]
	}

	@staticmethod
	def __get_browser_with_selenium(url):
		driver = webdriver.Chrome()
		driver.get(url)
		return driver

	@staticmethod
	def __get_page_with_selenium(driver):
		requiredHtml = driver.page_source
		soup = BeautifulSoup(requiredHtml, 'html.parser')
		liga_container = soup.find_all('div', {'data-name': 'dashboard-champ-content'})
		return liga_container

	@staticmethod
	def __get_container_with_coefficient(cont_coefficent):
		try:
			w1 = cont_coefficent.find('span', title='П1').find('span', class_='c-bets__inner').text
			draw = cont_coefficent.find('span', title='Ничья').find('span', class_='c-bets__inner').text
			w2 = cont_coefficent.find('span', title='П2').find('span', class_='c-bets__inner').text
		except AttributeError:
			w1 = '-'
			draw = '-'
			w2 = '-'
		try:
			total = cont_coefficent.find('span', class_='c-bets__bet non c-bets__bet_sm static-event num').text
			mtotal = cont_coefficent.find('span', title='Тотал больше').find('span', class_='c-bets__inner').text
			ltotal = cont_coefficent.find('span', title='Тотал меньше').find('span', class_='c-bets__inner').text
		except AttributeError:
			total = '-'
			mtotal = '-'
			ltotal = '-'
		return w1, draw, w2, total, mtotal, ltotal

	def get_data_from_line(self):
		url = 'https://1xstavka.ru/line/handball'
		driver = self.__get_browser_with_selenium(url)
		liga_container = self.__get_page_with_selenium(driver)
		for liga in liga_container:
			liga_name = liga.find('a', 'c-events__liga').text
			containers = liga.find_all('div', 'c-events__item c-events__item_col')	
			for container in containers:
				time_data = container.find('div', class_='c-events__time min').text
				name = container.find('span', class_='c-events__teams')['title']
				cofs_container = container.find('div', class_='c-bets')
				w1, draw, w2, total, mtotal, ltotal =\
								 self.__get_container_with_coefficient(cofs_container)
				data_parser_copy = copy.deepcopy(self.data_parser)

				data_parser_copy['time_data'].append(time_data)
				data_parser_copy['name'].append(name)
				data_parser_copy['liga_name'].append(liga_name)
				data_parser_copy['w1'].append(w1)
				data_parser_copy['draw'].append(draw)
				data_parser_copy['w2'].append(w2)
				data_parser_copy['total'].append(total)
				data_parser_copy['mtotal'].append(mtotal)
				data_parser_copy['ltotal'].append(ltotal)

		self.save_data_to_excel_line(data_parser_copy)
		driver.close()

	def get_data_from_live(self):
		url = 'https://1xstavka.ru/live/handball'
		driver = self.__get_browser_with_selenium(url)
		while True:
			liga_container = self.__get_page_with_selenium(driver)
			for liga in liga_container:
				liga_name = liga.find('a', 'c-events__liga').text
				containers = liga.find_all('div', 'c-events__item c-events__item_col')
				for container in containers:
					name = container.find('span', class_='c-events__teams')['title']
					status = container.find('span', class_='c-events__overtime').text
					time_data = container.find('div', class_='c-events__time').find('span').text
					cofs_container = container.find('div', class_='c-bets')
					w1, draw, w2, total, mtotal, ltotal =\
										 self.__get_container_with_coefficient(cofs_container)
					data_parser_copy = copy.deepcopy(self.data_parser)
					data_parser_copy['w1'].append(w1)
					data_parser_copy['draw'].append(draw)
					data_parser_copy['w2'].append(w2)
					data_parser_copy['name'].append(name)
					data_parser_copy['total_live'].append(total)
					data_parser_copy['mtotal_live'].append(mtotal)
					data_parser_copy['ltotal_live'].append(ltotal)

			self.save_data_to_excel_live(data_parser_copy)


if __name__ == '__main__':
	pt = ParserData()
	pt.get_data_from_line()
	pt.get_data_from_live()
	